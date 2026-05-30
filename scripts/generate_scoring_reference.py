"""
Generate a reference Excel workbook documenting the scoring/ranking logic.
Shows all rules, weights, formulas, lookup tables, and example calculations.
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "processed")
wb = Workbook()

# Styles
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
section_font = Font(bold=True, size=11, color="1A5276")
formula_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
hot_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
warm_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
nurture_fill = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid")
cold_fill = PatternFill(start_color="EAECEE", end_color="EAECEE", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)


# =============================================================================
# SHEET 1: Model Overview
# =============================================================================
ws = wb.active
ws.title = "Model Overview"

ws["A1"] = "LEAD/CONTACT READINESS SCORING MODEL"
ws["A1"].font = Font(bold=True, size=16)
ws["A3"] = "Purpose"
ws["A3"].font = section_font
ws["B3"] = "Rank every CRM record (lead or contact) on a 0-100 readiness scale"
ws["A4"] = "Question Answered"
ws["A4"].font = section_font
ws["B4"] = "Is this person worth a phone call RIGHT NOW?"
ws["A5"] = "Replaces"
ws["A5"].font = section_font
ws["B5"] = "Legacy MQL flag (binary Marketo threshold at 50 points)"

ws["A7"] = "COMPOSITE FORMULA"
ws["A7"].font = Font(bold=True, size=14)
ws["A8"] = "readiness_score ="
ws["A8"].font = Font(bold=True, size=11)
ws["B8"] = "(Engagement × 0.40) + (Profile × 0.25) + (Account × 0.20) + (Momentum × 0.15)"
ws["B8"].fill = formula_fill

ws["A10"] = "COMPONENT WEIGHTS"
ws["A10"].font = section_font
data = [
    ["Component", "Weight", "What It Measures", "Primary Signal"],
    ["Engagement Recency", "40%", "Recent real engagement activity", "Days since last response"],
    ["Profile Fit", "25%", "ICP alignment (seniority + persona)", "Job level × persona match"],
    ["Account Strength", "20%", "Company-level buying signals", "Named account + intent score"],
    ["Behavioral Momentum", "15%", "Is engagement accelerating?", "Last 30d vs prior 30d volume"],
]
for i, row in enumerate(data, start=11):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=val).border = thin_border
style_header(ws, 11, 4)

ws["A18"] = "TIER DEFINITIONS"
ws["A18"].font = section_font
tiers = [
    ["Tier", "Score Range", "BDR Action", "Expected %"],
    ["Hot", "70 - 100", "Call this week", "~5%"],
    ["Warm", "45 - 69", "Prioritize outreach", "~25%"],
    ["Nurture", "20 - 44", "Marketing nurture programs", "~50%"],
    ["Cold", "0 - 19", "Deprioritize / monitor", "~20%"],
]
fills = [None, hot_fill, warm_fill, nurture_fill, cold_fill]
for i, row in enumerate(tiers, start=19):
    for j, val in enumerate(row, start=1):
        cell = ws.cell(row=i, column=j, value=val)
        cell.border = thin_border
        if i > 19 and fills[i - 19]:
            cell.fill = fills[i - 19]
style_header(ws, 19, 4)

ws["A26"] = "KEY CONSTRAINTS"
ws["A26"].font = section_font
constraints = [
    "1. MQL status is NEVER an input (avoids circular dependency)",
    "2. Exclusions (competitor, opted-out, bounced) are orthogonal flags — NOT score penalties",
    "3. Only is_responded=True engagements count (automation discounting)",
    "4. Leads get 15-point account baseline (entity-type fairness)",
    "5. Score is always bounded 0-100 per component and in aggregate",
]
for i, c in enumerate(constraints, start=27):
    ws.cell(row=i, column=1, value=c)

auto_width(ws)

# =============================================================================
# SHEET 2: Engagement Scoring Rules
# =============================================================================
ws2 = wb.create_sheet("Engagement Rules")

ws2["A1"] = "ENGAGEMENT RECENCY SCORING (Weight: 40%)"
ws2["A1"].font = Font(bold=True, size=14)

ws2["A3"] = "TIME-DECAY FUNCTION"
ws2["A3"].font = section_font
ws2["A4"] = "Formula"
ws2["B4"] = "recency_score = 100 × EXP(-0.693 × days_since_last / 45)"
ws2["B4"].fill = formula_fill
ws2["A5"] = "Half-life"
ws2["B5"] = "45 days (score halves every 45 days without new engagement)"

decay_data = [
    ["Days Since Last Engagement", "Recency Score", "Interpretation"],
    [0, 100.0, "Engaged today"],
    [7, 89.7, "This week"],
    [15, 79.4, "Two weeks ago"],
    [30, 63.0, "One month ago"],
    [45, 50.0, "Half-life reached"],
    [60, 39.7, "Two months ago"],
    [90, 25.0, "One quarter ago"],
    [120, 15.8, "Four months ago"],
    [180, 6.3, "Six months ago"],
    [365, 0.4, "One year ago — essentially zero"],
]
for i, row in enumerate(decay_data, start=7):
    for j, val in enumerate(row, start=1):
        ws2.cell(row=i, column=j, value=val).border = thin_border
style_header(ws2, 7, 3)

ws2["A20"] = "ENGAGEMENT SCORING FORMULA"
ws2["A20"].font = section_font
ws2["A21"] = (
    "raw_score = (recency × 0.4) + (volume × 0.3) + diversity_bonus + high_value_bonus - automation_penalty"
)
ws2["A21"].fill = formula_fill

components = [
    ["Sub-Component", "Formula", "Cap", "Notes"],
    [
        "Recency",
        "100 × EXP(-0.693 × days / 45)",
        "100",
        "Exponential decay from last REAL engagement",
    ],
    [
        "Volume",
        "(engagements_30d × 20) + (engagements_90d × 5)",
        "100",
        "Diminishing returns via cap",
    ],
    ["Diversity Bonus", "unique_campaign_types × 10", "30", "Rewards multi-channel engagement"],
    [
        "High-Value Bonus",
        "(webinars × 15) + (events × 20) + (content × 10)",
        "40",
        "Premium engagement types",
    ],
    ["Automation Penalty", "automation_ratio × 30", "30", "Deducted; penalizes email-only sends"],
]
for i, row in enumerate(components, start=23):
    for j, val in enumerate(row, start=1):
        ws2.cell(row=i, column=j, value=val).border = thin_border
style_header(ws2, 23, 4)

ws2["A31"] = "WHAT COUNTS AS 'REAL' ENGAGEMENT"
ws2["A31"].font = section_font
engage_rules = [
    ["Campaign Type", "Counts (is_responded=True)", "Does NOT Count"],
    ["Webinar", "Attended", "Registered only, No Show"],
    ["Event", "Attended", "Registered only, No Show"],
    ["Content Syndication", "Downloaded, Responded", "—"],
    ["Email", "Opened, Clicked", "Sent (automation)"],
    ["Advertisement", "Clicked, Converted", "Impression only"],
    ["Telemarketing", "Contacted, Interested", "Not Interested"],
]
for i, row in enumerate(engage_rules, start=32):
    for j, val in enumerate(row, start=1):
        ws2.cell(row=i, column=j, value=val).border = thin_border
style_header(ws2, 32, 3)

auto_width(ws2)

# =============================================================================
# SHEET 3: Profile Scoring Rules
# =============================================================================
ws3 = wb.create_sheet("Profile Rules")

ws3["A1"] = "PROFILE FIT SCORING (Weight: 25%)"
ws3["A1"].font = Font(bold=True, size=14)

ws3["A3"] = "Formula"
ws3["B3"] = "profile_score = (level_score × 50) + (persona_score × 50) + completeness_bonus"
ws3["B3"].fill = formula_fill
ws3["A4"] = "Completeness"
ws3["B4"] = "+10 if title is populated, +10 if persona is populated"

ws3["A6"] = "JOB LEVEL LOOKUP"
ws3["A6"].font = section_font
levels = [
    ["Job Level", "Score (0-1)", "Points Contributed (×50)"],
    ["C-Level (CISO, CTO, CEO)", 1.0, 50],
    ["VP", 0.85, 42.5],
    ["Director", 0.7, 35],
    ["Manager", 0.5, 25],
    ["Individual Contributor", 0.3, 15],
    ["Unknown / NULL", 0.2, 10],
    ["Non-Prospect", 0.0, 0],
]
for i, row in enumerate(levels, start=7):
    for j, val in enumerate(row, start=1):
        ws3.cell(row=i, column=j, value=val).border = thin_border
style_header(ws3, 7, 3)

ws3["A17"] = "JOB PERSONA LOOKUP"
ws3["A17"].font = section_font
personas = [
    ["Job Persona", "Score (0-1)", "Points Contributed (×50)"],
    ["CISO", 1.0, 50],
    ["Technical Buyer", 0.9, 45],
    ["Financial Buyer", 0.7, 35],
    ["IT Operations", 0.6, 30],
    ["Security Engineer", 0.5, 25],
    ["Unknown / NULL", 0.2, 10],
    ["Non-Prospect", 0.0, 0],
]
for i, row in enumerate(personas, start=18):
    for j, val in enumerate(row, start=1):
        ws3.cell(row=i, column=j, value=val).border = thin_border
style_header(ws3, 18, 3)

ws3["A28"] = "EXAMPLE CALCULATIONS"
ws3["A28"].font = section_font
examples = [
    ["Record", "Level", "Persona", "Title?", "Persona?", "Score"],
    [
        "VP Security at ICP account",
        "VP (0.85)",
        "CISO (1.0)",
        "Yes (+10)",
        "Yes (+10)",
        "42.5 + 50 + 20 = 100 (capped)",
    ],
    [
        "Director of IT",
        "Director (0.7)",
        "Technical Buyer (0.9)",
        "Yes (+10)",
        "Yes (+10)",
        "35 + 45 + 20 = 100 (capped)",
    ],
    [
        "Unknown title, unknown persona",
        "NULL (0.2)",
        "NULL (0.2)",
        "No (+0)",
        "No (+0)",
        "10 + 10 + 0 = 20",
    ],
    [
        "Manager, Security Engineer",
        "Manager (0.5)",
        "Security Eng (0.5)",
        "Yes (+10)",
        "Yes (+10)",
        "25 + 25 + 20 = 70",
    ],
]
for i, row in enumerate(examples, start=29):
    for j, val in enumerate(row, start=1):
        ws3.cell(row=i, column=j, value=val).border = thin_border
style_header(ws3, 29, 6)

auto_width(ws3)

# =============================================================================
# SHEET 4: Account Scoring Rules
# =============================================================================
ws4 = wb.create_sheet("Account Rules")

ws4["A1"] = "ACCOUNT STRENGTH SCORING (Weight: 20%)"
ws4["A1"].font = Font(bold=True, size=14)

ws4["A3"] = "Formula (contacts with accounts)"
ws4["B3"] = "account_score = named(30) + icp(25) + intent(0-30) + size(0-15)"
ws4["B3"].fill = formula_fill
ws4["A4"] = "Formula (leads without accounts)"
ws4["B4"] = "account_score = 15 (fixed baseline for entity-type fairness)"
ws4["B4"].fill = formula_fill

ws4["A6"] = "COMPONENT BREAKDOWN"
ws4["A6"].font = section_font
acct_rules = [
    ["Signal", "Points", "Condition", "Rationale"],
    [
        "Named Account",
        "+30",
        "is_named_account = True",
        "Manually curated target list (~2000 companies)",
    ],
    [
        "ICP Qualified",
        "+25",
        "is_icp_qualified = True",
        "Firmographic match to ideal customer profile",
    ],
    ["Intent Score", "0-30", "intent_score / 100 × 30", "Third-party intent data (6sense/Bombora)"],
    ["Company Size", "0-15", "(employee_score + revenue_score) × 7.5", "Larger = more budget/need"],
    [
        "Lead Baseline",
        "15 (fixed)",
        "Entity has no account_id",
        "Prevents unfair lead disadvantage",
    ],
]
for i, row in enumerate(acct_rules, start=7):
    for j, val in enumerate(row, start=1):
        ws4.cell(row=i, column=j, value=val).border = thin_border
style_header(ws4, 7, 4)

ws4["A15"] = "SIZE SCORING"
ws4["A15"].font = section_font
size_rules = [
    ["Employee Count", "employee_score", "Annual Revenue", "revenue_score"],
    ["50", "0.005", "$1M", "0.001"],
    ["500", "0.05", "$10M", "0.01"],
    ["1,000", "0.10", "$50M", "0.05"],
    ["5,000", "0.50", "$500M", "0.50"],
    ["10,000+", "1.00 (capped)", "$1B+", "1.00 (capped)"],
]
for i, row in enumerate(size_rules, start=16):
    for j, val in enumerate(row, start=1):
        ws4.cell(row=i, column=j, value=val).border = thin_border
style_header(ws4, 16, 4)

auto_width(ws4)

# =============================================================================
# SHEET 5: Momentum Scoring Rules
# =============================================================================
ws5 = wb.create_sheet("Momentum Rules")

ws5["A1"] = "BEHAVIORAL MOMENTUM SCORING (Weight: 15%)"
ws5["A1"].font = Font(bold=True, size=14)

ws5["A3"] = "Formula"
ws5["B3"] = "IF engagements_30d = 0 THEN 0 ELSE MIN(momentum × 40, 80) + IF(accelerating, 20, 0)"
ws5["B3"].fill = formula_fill
ws5["A4"] = "Momentum calc"
ws5["B4"] = "momentum = (last_30d_count - prior_30d_count) / MAX(prior_30d_count, 1)"
ws5["B4"].fill = formula_fill

ws5["A6"] = "DECISION TABLE"
ws5["A6"].font = section_font
momentum_rules = [
    ["Last 30d", "Prior 30d", "Momentum", "Accelerating?", "Score"],
    [0, "any", "N/A", "N/A", "0 (no recent activity)"],
    [3, 0, "1.0 (new burst)", "Yes", "MIN(40, 80) + 20 = 60"],
    [5, 2, "1.5", "Yes", "MIN(60, 80) + 20 = 80"],
    [2, 2, "0.0", "No", "MIN(0, 80) + 0 = 0"],
    [1, 3, "-0.67 → 0", "No", "0 (decelerating)"],
    [4, 1, "3.0 → capped 2.0", "Yes", "MIN(80, 80) + 20 = 100"],
]
for i, row in enumerate(momentum_rules, start=7):
    for j, val in enumerate(row, start=1):
        ws5.cell(row=i, column=j, value=val).border = thin_border
style_header(ws5, 7, 5)

auto_width(ws5)

# =============================================================================
# SHEET 6: Exclusion Rules
# =============================================================================
ws6 = wb.create_sheet("Exclusion Rules")

ws6["A1"] = "EXCLUSION FLAGS (Orthogonal to Score)"
ws6["A1"].font = Font(bold=True, size=14)
ws6["A2"] = "These are NOT score penalties. A record can score 90/100 and still be excluded."

ws6["A4"] = "LEAD EXCLUSIONS"
ws6["A4"].font = section_font
lead_excl = [
    ["Flag", "Detection Rule", "Effect"],
    ["Opted Out", "has_opted_out = True", "Cannot email — legal/compliance block"],
    ["Email Bounced", "email_bounced = True", "No deliverable address"],
    ["Competitor", "company IN (CrowdStrike, Palo Alto, ...)", "Not a prospect"],
    ["Non-Prospect", "job_persona = 'Non-Prospect'", "Partner/vendor/employee"],
    ["Disqualified", "is_disqualified = True", "Previously rejected by BDR/AE"],
]
for i, row in enumerate(lead_excl, start=5):
    for j, val in enumerate(row, start=1):
        ws6.cell(row=i, column=j, value=val).border = thin_border
style_header(ws6, 5, 3)

ws6["A13"] = "CONTACT EXCLUSIONS"
ws6["A13"].font = section_font
contact_excl = [
    ["Flag", "Detection Rule", "Effect"],
    ["Opted Out", "has_opted_out = True", "Cannot email"],
    ["No Longer With Company", "no_longer_with_company = True", "Wrong person at account"],
    ["Non-Prospect", "job_persona = 'Non-Prospect'", "Partner/vendor/employee"],
    ["DNC Account", "account.do_not_contact = True", "Legal block at company level"],
]
for i, row in enumerate(contact_excl, start=14):
    for j, val in enumerate(row, start=1):
        ws6.cell(row=i, column=j, value=val).border = thin_border
style_header(ws6, 14, 3)

ws6["A21"] = (
    "IMPORTANT: Excluded records are removed from the callable list but retain their score."
)
ws6["A22"] = "This lets analysts ask: 'What would this competitor score if they were a prospect?'"

auto_width(ws6)

# =============================================================================
# SHEET 7: Worked Examples
# =============================================================================
ws7 = wb.create_sheet("Worked Examples")

ws7["A1"] = "WORKED EXAMPLES — Full Score Calculation"
ws7["A1"].font = Font(bold=True, size=14)

examples_full = [
    [
        "#",
        "Persona",
        "Eng Score",
        "Profile Score",
        "Account Score",
        "Momentum Score",
        "Readiness Score",
        "Tier",
        "Excluded?",
    ],
    [
        1,
        "VP Security, named ICP, 3 webinars this month",
        85,
        100,
        95,
        80,
        "85×.4+100×.25+95×.2+80×.15=90.0",
        "Hot",
        "No",
    ],
    [
        2,
        "Same profile, engagement >6 months old",
        8,
        100,
        95,
        0,
        "8×.4+100×.25+95×.2+0×.15=47.2",
        "Warm",
        "No",
    ],
    [
        3,
        "Junior analyst, 15 responses in 30 days",
        92,
        35,
        15,
        90,
        "92×.4+35×.25+15×.2+90×.15=62.1",
        "Warm",
        "No",
    ],
    [
        4,
        "CISO Fortune 500, zero engagement",
        0,
        100,
        85,
        0,
        "0×.4+100×.25+85×.2+0×.15=42.0",
        "Nurture",
        "No",
    ],
    [
        5,
        "Competitor, high engagement",
        88,
        40,
        15,
        60,
        "88×.4+40×.25+15×.2+60×.15=57.2",
        "Warm",
        "YES — Competitor",
    ],
    [
        6,
        "Bounced+opted out, recent event",
        70,
        45,
        15,
        50,
        "70×.4+45×.25+15×.2+50×.15=49.8",
        "Warm",
        "YES — Bounced+OptOut",
    ],
    [
        7,
        "Broken conversion link, split engagement",
        30,
        60,
        15,
        0,
        "30×.4+60×.25+15×.2+0×.15=30.0",
        "Nurture",
        "No — but DQ flagged",
    ],
    [
        8,
        "40 campaigns, 38 auto emails (DQ-8)",
        12,
        50,
        15,
        0,
        "12×.4+50×.25+15×.2+0×.15=20.3",
        "Nurture",
        "No",
    ],
    [
        9,
        "Re-MQL'd 4 times (cycling)",
        55,
        45,
        15,
        40,
        "55×.4+45×.25+15×.2+40×.15=42.3",
        "Nurture",
        "No",
    ],
    [
        10,
        "CC contact, high-intent named acct, 2 form fills",
        60,
        70,
        90,
        45,
        "60×.4+70×.25+90×.2+45×.15=66.3",
        "Warm",
        "No",
    ],
]
for i, row in enumerate(examples_full, start=3):
    for j, val in enumerate(row, start=1):
        cell = ws7.cell(row=i, column=j, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True)
style_header(ws7, 3, 9)

ws7["A15"] = "NOTE: Scores are approximate — actual values depend on exact feature calculations."
ws7["A16"] = (
    "The formula column shows: (Eng×0.4) + (Profile×0.25) + (Account×0.2) + (Momentum×0.15)"
)

auto_width(ws7)
ws7.column_dimensions["B"].width = 40
ws7.column_dimensions["G"].width = 35

# =============================================================================
# Save
# =============================================================================
output_path = os.path.join(OUTPUT_DIR, "scoring_model_reference.xlsx")
wb.save(output_path)
print(f"Saved: {output_path}")
